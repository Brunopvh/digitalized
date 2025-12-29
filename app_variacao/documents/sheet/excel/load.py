from __future__ import annotations
from io import BytesIO
from abc import ABC, abstractmethod
from openpyxl import load_workbook, Workbook
from typing import Any
import pandas as pd
import os.path
import zipfile
import xml.etree.ElementTree as ET

# Módulos locais
from app_variacao.documents.erros import *
from app_variacao.documents.sheet.types import SheetData, WorkbookData, SheetIndexNames
from app_variacao.documents.sheet.xml import read_zip_xml, WorkbookMappingXML
from app_variacao.documents.sheet.excel._col_index import column_coord_to_index
from app_variacao.types.core import ObjectAdapter
from app_variacao.util import get_md5_bytes


class ExcelLoad(ABC):

    @abstractmethod
    def hash(self) -> int:
        pass

    @abstractmethod
    def get_sheet_index(self) -> SheetIndexNames:
        pass

    @abstractmethod
    def get_workbook_data(self) -> WorkbookData:
        pass

    def get_sheet_at(self, idx: int) -> SheetData:
        idx_sheet_names: SheetIndexNames = self.get_sheet_index()
        name = idx_sheet_names[idx]
        return self.get_workbook_data()[name]

    def get_sheet(self, sheet_name: str | None) -> SheetData:
        if sheet_name is not None:
            return self.get_workbook_data()[sheet_name]
        return self.get_sheet_at(0)


class ExcelLoadOpenPyxl(ExcelLoad):

    def __init__(self, file_excel: str | BytesIO):
        super().__init__()
        self.file_excel: str | BytesIO = file_excel
        self.__workbook: Workbook | None = None
        self.__hash: int = hash(file_excel)

    def hash(self) -> int:
        return self.__hash

    def __get_wb(self) -> Workbook:
        if self.__workbook is None:
            try:
                self.__workbook = load_workbook(self.file_excel)
            except FileNotFoundError:
                raise LoadWorkbookError(f'{__class__.__name__} FileNotFoundError()')
            except Exception as e:
                raise LoadWorkbookError(f'{__class__.__name__} {e}')
        return self.__workbook

    def __get_sheet_data_from_name(self, name: str) -> SheetData:
        sheet_data = SheetData()
        wb: Workbook = self.__get_wb()
        ws = wb[name]  # Acessa a planilha pelo nome
        # Iterador de linhas
        rows_iterator = ws.rows

        # Extrai e processa o cabeçalho (primeira linha)
        try:
            # O 'next' avança o iterador para a segunda linha após esta chamada
            header_cells = next(rows_iterator)
        except StopIteration:
            # Se a planilha estiver vazia (sem cabeçalho), retorna vazio
            return sheet_data

        # Prepara o SheetData com as chaves do cabeçalho
        # É importante armazenar o cabeçalho como uma lista separada para indexação
        header_list: list[str] = []
        for cell in header_cells:
            # Garante que o nome da coluna seja uma string (usando str() no valor)
            col_name = str(cell.value) if cell.value is not None else ''
            # Adiciona o nome da coluna à lista de cabeçalho
            header_list.append(col_name)
            # Inicializa a lista de valores para cada coluna no SheetData
            sheet_data[col_name] = []

        # 3. Itera sobre as linhas restantes (dados)
        for row in rows_iterator:

            # Itera sobre as células da linha
            for col_index, cell in enumerate(row):
                # Garante que a célula tem um índice correspondente ao cabeçalho
                if col_index < len(header_list):
                    # Pega o nome da coluna (chave) usando o índice
                    col_name = header_list[col_index]
                    # Converte o valor da célula para string (conforme o requisito de SheetData)
                    cell_value_str = str(cell.value) if cell.value is not None else ''
                    # Adiciona o valor à lista da coluna correspondente no SheetData
                    sheet_data[col_name].append(cell_value_str)
        return sheet_data

    def get_sheet_index(self) -> SheetIndexNames:
        """

        :return: Dicionário dict/SheetIndexNames sendo as chaves os indíces (int) de
        cada planilha e os valores (str) os respectivos nomes.
        """
        wb = self.__get_wb()
        sheet_index = SheetIndexNames()
        names: list[str] = wb.sheetnames
        for num, name in enumerate(names):
            #sheet_index[num] = name
            sheet_index.add_index(num, name)
        return sheet_index

    def get_workbook_data(self) -> WorkbookData:
        final = WorkbookData()
        sheet_index = self.get_sheet_index()
        for sheet_name in sheet_index.get_sheet_names():
            sheet_data = self.__get_sheet_data_from_name(sheet_name)
            final.add_sheet(sheet_name, sheet_data)
        return final


class ExcelLoadPandas(ExcelLoad):

    def __init__(self, xlsx_file):
        self.xlsx_file: str | BytesIO = xlsx_file
        self.__hash: int = hash(xlsx_file)

    def hash(self) -> int:
        return self.__hash

    def get_sheet_index(self) -> SheetIndexNames:
        rd: pd.ExcelFile = pd.ExcelFile(self.xlsx_file)
        names = [str(x) for x in rd.sheet_names]
        return SheetIndexNames.create_from_list(names)

    def get_workbook_data(self) -> WorkbookData:
        data: dict[Any, pd.DataFrame] = pd.read_excel(self.xlsx_file, sheet_name=None)
        workbook_data = WorkbookData()
        for _key in data.keys():
            df: pd.DataFrame = data[_key]
            workbook_data.add_sheet(str(_key), SheetData.create_from_data(df))
        return workbook_data

    def get_sheet_at(self, idx: int) -> SheetData:
        return super().get_sheet_at(idx)

    def get_sheet(self, sheet_name: str | None) -> SheetData:
        return super().get_sheet(sheet_name)

#===========================================================#
# Classe para mapear os arquivos XML compactados no .xlsx
#===========================================================#


class XlsxMapXMLFiles(dict[str, str]):
    """Mapeia nomes lógicos para caminhos internos do ZIP."""

    def __init__(self):
        super().__init__({})
        self['workbook'] = os.path.join('xl', 'workbook.xml')
        self['shared_strings'] = os.path.join('xl', 'sharedStrings.xml')
        self['dir_work_sheets'] = os.path.join('xl', 'worksheets')

    def get_xml_workbook(self) -> str:
        return self['workbook']

    def get_xml_shared_strings(self) -> str:
        return self['shared_strings']

    def get_sheet_path(self, sheet_xml_prefix: str) -> str:
        """
        :param sheet_xml_prefix: Ex: sheet1, sheet2
        :return: caminho do xml sheet. Ex: xl/worksheets/sheet1.xml
        """
        return os.path.join(self['dir_work_sheets'], f'{sheet_xml_prefix}.xml')

#===========================================================#
# Classe de Leitura XML
#===========================================================#


class XMLProcessor(object):
    """
    Contém a lógica de baixo nível para processar XMLs internos, de arquivos
    .xlsx
    """

    def __init__(self, xlsx_input: str | BytesIO):
        self.xlsx_input = xlsx_input
        self.xml_file_names = XlsxMapXMLFiles()
        self.shared_strings: list[str] = []
        self.workbook_map: WorkbookMappingXML | None = None

    def load_shared_strings(self, zf: zipfile.ZipFile):
        """Carrega strings compartilhadas de 'xl/sharedStrings.xml'."""
        self.shared_strings = []  # Limpa strings anteriores

        tree: ET.ElementTree
        try:
            tree, err = read_zip_xml(zf, self.xml_file_names.get_xml_shared_strings())
            if tree is None:
                return

            root = tree.getroot()
            ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            for si in root.findall('s:si', ns):
                text_element = si.find('s:t', ns)
                if text_element is not None:
                    self.shared_strings.append(text_element.text or '')
                else:
                    rich_text = ''.join(t.text for t in si.findall('s:r/s:t', ns) if t.text)
                    self.shared_strings.append(rich_text or '')
        except Exception:
            pass  # Mantém lista vazia em caso de erro

    def load_workbook_mapping(self, zf: zipfile.ZipFile) -> WorkbookMappingXML:
        """Mapeia nomes de abas a seus respectivos arquivos sheet*.xml."""
        __map = WorkbookMappingXML()
        tree: ET.ElementTree
        try:
            tree, err = read_zip_xml(zf, self.xml_file_names.get_xml_workbook())
            if tree is None:
                return __map

            root = tree.getroot()
            ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            sheets: ET.Element | None = root.find('s:sheets', ns)
            if sheets is not None:
                for sheet in sheets.findall('s:sheet', ns):
                    _sheet_name: str | None = sheet.get('name')
                    sheet_id: str | None = sheet.get('sheetId')
                    #rel_id = sheet.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')

                    if _sheet_name and sheet_id:
                        # Assumimos 'sheet{sheet_id}' como prefixo XML.
                        # Em um cenário 100% robusto, usaria-se rel_id para buscar o nome exato no .rels.
                        sheet_xml_prefix = f'sheet{sheet_id}'
                        __map.set_sheet_id_and_prefix(_sheet_name, sheet_id, sheet_xml_prefix)
        except Exception:
            pass
        return __map

    def process_sheet_to_sheetdata(self, sheet_name: str) -> SheetData:
        """
        Lê e processa o XML de uma aba específica, retornando um objeto SheetData.
        Esta é a parte que adapta a leitura XML ao formato SheetData (coluna-orientado).
        """
        sheet_data = SheetData()

        if self.workbook_map is None:
            raise Exception("Workbook mapping não carregado.")  # Deve ser carregado no __get_workbook_data

        sheet_id = self.workbook_map.get_sheet_id_from_name(sheet_name)
        if sheet_id is None:
            return sheet_data  # Nome da aba não encontrado

        xml_prefix = self.workbook_map.get_xml_sheet_prefix_from_id(sheet_id)
        sheet_path = self.xml_file_names.get_sheet_path(xml_prefix)

        # 1. Armazenamento temporário de linhas: lista de dicionários {col_idx: valor}
        rows_map: dict[int, dict[int, Any]] = {}
        max_col_idx = 0
        max_row_idx = 0

        try:
            with zipfile.ZipFile(self.xlsx_input, 'r') as zf:

                # Garante que as strings compartilhadas estão carregadas
                if not self.shared_strings:
                    self.load_shared_strings(zf)
                tree: ET.ElementTree
                tree, err = read_zip_xml(zf, sheet_path)
                if tree is None:
                    return sheet_data

                root = tree.getroot()
                ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

                sheet_data_element = root.find('s:sheetData', ns)
                if sheet_data_element is None:
                    return sheet_data

                for row_element in sheet_data_element.findall('s:row', ns):
                    # Número da linha (base 1)
                    row_num: int = int(row_element.get('r', 1))
                    rows_map[row_num - 1] = {}  # Mapeia para índice base 0
                    max_row_idx = max(max_row_idx, row_num - 1)

                    for cell_element in row_element.findall('s:c', ns):
                        sheet_coord: str | None = cell_element.get('r')
                        if not sheet_coord:
                            continue

                        col_idx_base1: int = column_coord_to_index(sheet_coord)
                        col_idx_base0: int = col_idx_base1 - 1

                        max_col_idx = max(max_col_idx, col_idx_base0)

                        cell_type: str | None = cell_element.get('t')
                        v_element: ET.Element | None = cell_element.find('s:v', ns)
                        element_value: str = v_element.text if v_element is not None else ''

                        #final_value: Any = None

                        if cell_type == 's' and element_value.isdigit():
                            try:
                                final_value = self.shared_strings[int(element_value)]
                            except (ValueError, IndexError):
                                final_value = ''

                        elif cell_type == 'n' or cell_type is None:
                            try:
                                final_value = float(element_value)
                                if final_value == int(final_value):
                                    final_value = int(final_value)
                            except ValueError:
                                final_value = element_value if element_value else None

                        elif element_value:
                            final_value = element_value
                        else:
                            final_value = None

                        rows_map[row_num - 1][col_idx_base0] = str(final_value or '')  # Armazena como string

        except Exception as e:
            print(f"{__class__.__name__} Erro ao processar XML da aba: {e}")
            return sheet_data

        # 2. Reorganizar de Linha-Orientado para Coluna-Orientado (SheetData)
        if not rows_map:
            return sheet_data

        # 2.1. Extrair Cabeçalho e determinar o número total de colunas
        header_row = rows_map.get(0, {})  # Cabeçalho é a primeira linha (índice 0)

        # Cria uma lista de cabeçalhos padronizada (usando A, B, C... se o nome não for dado)
        header_list: list[str] = [
            header_row.get(i, f"Col{i + 1}")  # Usa "ColX" se a célula do cabeçalho estiver vazia
            for i in range(max_col_idx + 1)
        ]

        # Inicializa SheetData com as colunas
        for col_name in header_list:
            sheet_data.add_column(col_name, [])

        # 2.2. Preencher os dados
        for row_idx in range(1, max_row_idx + 1):  # Começa da linha 1 (dados)
            current_row = rows_map.get(row_idx, {})

            for col_idx, col_name in enumerate(header_list):
                # Pega o valor se existir, senão usa string vazia
                value = current_row.get(col_idx, '')

                # Adiciona o valor à coluna correspondente no SheetData
                sheet_data[col_name].append(str(value))

        return sheet_data

#===========================================================#
# Classe de Leitura XML Implementa ExcelLoad()
#===========================================================#


class ExcelLoadXML(ExcelLoad):
    """
    Carregador de Excel que lê o arquivo .xlsx diretamente como um arquivo ZIP
    contendo XMLs (padrão Open XML).
    """

    def __init__(self, file_excel: str | BytesIO):
        super().__init__()
        self.file_excel: str | BytesIO = file_excel
        self.__processor: XMLProcessor = XMLProcessor(file_excel)
        self.__workbook_data: WorkbookData | None = None
        self.__sheet_index_names: SheetIndexNames | None = None
        self.__hash: int = hash(file_excel)

    def hash(self) -> int:
        return self.__hash

    def get_sheet_index(self) -> SheetIndexNames:
        if self.__sheet_index_names is None:
            # Garante que o WorkbookMapping foi carregado
            if self.__processor.workbook_map is None:
                _ = self.get_workbook_data()

            # Cria SheetIndexNames a partir dos nomes mapeados
            if self.__processor.workbook_map:
                names = self.__processor.workbook_map.get_sheet_names()
                self.__sheet_index_names = SheetIndexNames.create_from_list(names)
            else:
                self.__sheet_index_names = SheetIndexNames()  # Vazio

        return self.__sheet_index_names

    def get_workbook_data(self) -> WorkbookData:
        if self.__workbook_data is None:

            # 1. Carregar mapeamento do workbook
            try:
                with zipfile.ZipFile(self.file_excel, 'r') as zf:
                    # Carrega shared strings e mapeamento
                    self.__processor.load_shared_strings(zf)
                    self.__processor.workbook_map = self.__processor.load_workbook_mapping(zf)
            except Exception as e:
                # print(f"Erro ao carregar mapeamento XML: {e}")
                raise LoadWorkbookError(f'ExcelLoadXML: Erro ao carregar arquivo ZIP/XML: {e}')

            workbook_data = WorkbookData()
            if self.__processor.workbook_map:

                # 2. Iterar sobre as abas e processar uma por uma
                for sheet_name in self.__processor.workbook_map.get_sheet_names():
                    sheet_data = self.__processor.process_sheet_to_sheetdata(sheet_name)
                    workbook_data.add_sheet(sheet_name, sheet_data)

            self.__workbook_data = workbook_data
        return self.__workbook_data

#===========================================================#
# Classe Adapter para leitura de planihas Excel/Xlsx
#===========================================================#


class ReadSheetExcel(ObjectAdapter):

    def __init__(self, reader: ExcelLoad):
        super().__init__()
        self.__reader: ExcelLoad = reader

    def get_implementation(self) -> ExcelLoad:
        return self.__reader

    def hash(self) -> int:
        return self.get_implementation().hash()

    def get_workbook_data(self) -> WorkbookData:
        return self.__reader.get_workbook_data()

    def get_sheet_at(self, idx: int) -> SheetData:
        return self.__reader.get_sheet_at(idx)

    def get_sheet(self, sheet_name: str | None = None) -> SheetData:
        return self.__reader.get_sheet(sheet_name)

    def get_sheet_index(self) -> SheetIndexNames:
        return self.__reader.get_sheet_index()

    @classmethod
    def create_load_open_pyxl(cls, file_excel: str | BytesIO) -> ReadSheetExcel:
        rd = ExcelLoadOpenPyxl(file_excel)
        return cls(rd)

    @classmethod
    def create_load_pandas(cls, file_excel: str | BytesIO) -> ReadSheetExcel:
        rd = ExcelLoadPandas(file_excel)
        return cls(rd)

    @classmethod
    def create_load_xml(cls, file_excel: str | BytesIO) -> ReadSheetExcel:
        """Cria uma instância usando o carregador XML (do zero)."""
        rd = ExcelLoadXML(file_excel)
        return cls(rd)


__all__ = ['ReadSheetExcel', 'ExcelLoad']
