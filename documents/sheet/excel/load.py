from __future__ import annotations
from io import BytesIO
from abc import ABC, abstractmethod
from openpyxl import load_workbook, Workbook
from collections.abc import Iterator
import pandas as pd
from documents.erros import *
from typing import Any

class RowIterator(Iterator):

    def __init__(self, sheet: SheetData, reverse: bool = False):
        self.sheet: SheetData = sheet
        self.reverse: bool = reverse

        if self.reverse:
            self.__current_idx: int = -1
            self.__max_idx: int = -(len(self.sheet.get_first())) - 1
        else:
            self.__current_idx: int = 0
            self.__max_idx: int = len(self.sheet.get_first())

    def has_next(self) -> bool:
        if self.reverse:
            if self.__current_idx <= self.__max_idx:
                return False
        else:
            if self.__current_idx >= self.__max_idx:
                return False
        return True

    def __next__(self) -> list[str]:
        if not self.has_next():
            raise StopIteration()
        idx_value = self.__current_idx
        if self.reverse:
            self.__current_idx -= 1
        else:
            self.__current_idx += 1
        return self.sheet.get_row(idx_value)


class SheetIndexNames(dict[int, str]):

    def __init__(self):
        super().__init__({})

    def keys(self) -> list[int]:
        return list(super().keys())

    def values(self) -> list[str]:
        return list(super().values())

    def add_index(self, idx: int, name: str):
        self[idx] = name

    def get_index_from_name(self, sheet_name: str) -> int | None:
        """
        Retorna o indíce de uma planilha a partir do nome.
        :param sheet_name: string nome da planilha
        :return: int indíce de uma planilha
        """
        for idx in self.keys():
            if self[idx] == sheet_name:
                return idx
        return None

    def get_sheet_name_at(self, idx: int) -> str:
        return self[idx]

    def get_sheet_names(self) -> list[str]:
        return self.values()

    @classmethod
    def create_from_list(cls, sheet_names: list[str]):
        sheet_idx = cls()
        for n, name in enumerate(sheet_names):
            sheet_idx.add_index(n, name)
        return sheet_idx


class SheetData(dict[str, list[str]]):

    def __init__(self):
        super().__init__({})

    def get_row(self, idx: int) -> list[str]:
        row = list()
        for _k in self.keys():
            value = self[_k][idx]
            row.append(value)
        return row

    def row_iterator(self, reverse: bool = False) -> RowIterator:
        return RowIterator(self, reverse)

    def get_first(self) -> list[str]:
        _k = self.keys()[0]
        return self[_k]

    def get_last(self) -> list[str]:
        _k = self.keys()[-1]
        return self[_k]

    def header(self) -> list[str]:
        return self.keys()

    def keys(self) -> list[str]:
        return list(super().keys())

    def values(self) -> list[list[str]]:
        return list(super().values())

    def add_column(self, head: str, column: list[str]):
        self[head] = column

    def get_max_rows(self) -> int:
        _keys = self.keys()
        _name = _keys[0]
        return len(self[_name])

    def to_data_frame(self) -> pd.DataFrame:
        return pd.DataFrame(self)

    @classmethod
    def create_from_data(cls, data: pd.DataFrame) -> SheetData:
        columns: list[str] = data.astype('str').columns.to_list()
        sheet_data = cls()
        for col in columns:
            sheet_data.add_column(col, data[col].astype('str').values.tolist())
        return sheet_data


class WorkbookData(dict[str, SheetData]):

    def __init__(self):
        super().__init__({})
        self.__sheet_index_names: SheetIndexNames | None = None

    def get_sheet_index_names(self) -> SheetIndexNames:
        """
        Dicionário com os nomes das planilhas apontando para os respetivos indices.
        :return: Dicionário com int, str
        """
        if self.__sheet_index_names is None:
            self.__sheet_index_names = SheetIndexNames()
            for n, name in enumerate(self.keys()):
                self.__sheet_index_names.add_index(n, name)
        return self.__sheet_index_names

    def set_sheet_index_names(self, sheet_index_names: SheetIndexNames):
        self.__sheet_index_names = sheet_index_names

    def get_first(self) -> SheetData:
        return self.get_sheet_at(0)

    def get_last(self) -> SheetData:
        _keys = self.keys()
        _last_key = _keys[-1]
        return self[_last_key]

    def get_sheet_at(self, idx: int) -> SheetData:
        _keys = self.keys()
        return self[_keys[idx]]

    def keys(self) -> list[str]:
        return list(super().keys())

    def values(self) -> list[SheetData]:
        return list(super().values())

    def add_sheet(self, name: str, sheet: SheetData):
        self[name] = sheet



class ExcelLoad(ABC):

    @abstractmethod
    def get_sheet_index(self) -> SheetIndexNames:
        pass

    @abstractmethod
    def get_workbook_data(self) -> WorkbookData:
        pass

    def get_sheet_at(self, idx: int) -> SheetData:
        idx_sheet_names: SheetIndexNames = self.get_sheet_index()
        name = idx_sheet_names[idx]
        return self.get_workbook_data()[name]

    def get_sheet(self, sheet_name: str | None) -> SheetData:
        if sheet_name is not None:
            return self.get_workbook_data()[sheet_name]
        return self.get_sheet_at(0)


class ExcelLoadOpenPyxl(ExcelLoad):

    def __init__(self, file_excel: str | BytesIO):
        super().__init__()
        self.file_excel: str | BytesIO = file_excel
        self.__workbook: Workbook | None = None

    def __get_wb(self) -> Workbook:
        if self.__workbook is None:
            try:
                self.__workbook = load_workbook(self.file_excel)
            except FileNotFoundError:
                raise LoadWorkbookError(f'{__class__.__name__} FileNotFoundError()')
            except Exception as e:
                raise LoadWorkbookError(f'{__class__.__name__} {e}')
        return self.__workbook

    def __get_sheet_data_from_name(self, name: str) -> SheetData:
        sheet_data = SheetData()
        wb: Workbook = self.__get_wb()
        ws = wb[name]  # Acessa a planilha pelo nome
        # Iterador de linhas
        rows_iterator = ws.rows

        # Extrai e processa o cabeçalho (primeira linha)
        try:
            # O 'next' avança o iterador para a segunda linha após esta chamada
            header_cells = next(rows_iterator)
        except StopIteration:
            # Se a planilha estiver vazia (sem cabeçalho), retorna vazio
            return sheet_data

        # Prepara o SheetData com as chaves do cabeçalho
        # É importante armazenar o cabeçalho como uma lista separada para indexação
        header_list: list[str] = []
        for cell in header_cells:
            # Garante que o nome da coluna seja uma string (usando str() no valor)
            col_name = str(cell.value) if cell.value is not None else ''
            # Adiciona o nome da coluna à lista de cabeçalho
            header_list.append(col_name)
            # Inicializa a lista de valores para cada coluna no SheetData
            sheet_data[col_name] = []

        # 3. Itera sobre as linhas restantes (dados)
        for row in rows_iterator:

            # Itera sobre as células da linha
            for col_index, cell in enumerate(row):
                # Garante que a célula tem um índice correspondente ao cabeçalho
                if col_index < len(header_list):
                    # Pega o nome da coluna (chave) usando o índice
                    col_name = header_list[col_index]
                    # Converte o valor da célula para string (conforme o requisito de SheetData)
                    cell_value_str = str(cell.value) if cell.value is not None else ''
                    # Adiciona o valor à lista da coluna correspondente no SheetData
                    sheet_data[col_name].append(cell_value_str)
        return sheet_data

    def get_sheet_index(self) -> SheetIndexNames:
        """

        :return: Dicionário dict/SheetIndexNames sendo as chaves os indíces (int) de
        cada planilha e os valores (str) os respectivos nomes.
        """
        wb = self.__get_wb()
        sheet_index = SheetIndexNames()
        names: list[str] = wb.sheetnames
        for num, name in enumerate(names):
            #sheet_index[num] = name
            sheet_index.add_index(num, name)
        return sheet_index

    def get_workbook_data(self) -> WorkbookData:
        final = WorkbookData()
        sheet_index = self.get_sheet_index()
        for sheet_name in sheet_index.get_sheet_names():
            sheet_data = self.__get_sheet_data_from_name(sheet_name)
            final.add_sheet(sheet_name, sheet_data)
        return final


class ExcelLoadPandas(ExcelLoad):

    def __init__(self, xlsx_file):
        self.xlsx_file: str | BytesIO = xlsx_file

    def get_sheet_index(self) -> SheetIndexNames:
        rd: pd.ExcelFile = pd.ExcelFile(self.xlsx_file)
        names = [str(x) for x in rd.sheet_names]
        return SheetIndexNames.create_from_list(names)

    def get_workbook_data(self) -> WorkbookData:
        data: dict[Any, pd.DataFrame] = pd.read_excel(self.xlsx_file, sheet_name=None)
        workbook_data = WorkbookData()
        for _key in data.keys():
            df: pd.DataFrame = data[_key]
            workbook_data.add_sheet(str(_key), SheetData.create_from_data(df))
        return workbook_data

    def get_sheet_at(self, idx: int) -> SheetData:
        return super().get_sheet_at(idx)

    def get_sheet(self, sheet_name: str | None) -> SheetData:
        return super().get_sheet(sheet_name)


class ReadSheetExcel(object):

    def __init__(self, reader: ExcelLoad):
        self.reader: ExcelLoad = reader

    def get_workbook_data(self) -> WorkbookData:
        return self.reader.get_workbook_data()

    def get_sheet_at(self, idx: int) -> SheetData:
        return self.reader.get_sheet_at(idx)

    def get_sheet(self, sheet_name: str | None = None) -> SheetData:
        return self.reader.get_sheet(sheet_name)

    def get_sheet_index(self) -> SheetIndexNames:
        return self.reader.get_sheet_index()

    @classmethod
    def create_load_open_pyxl(cls, file_excel: str | BytesIO) -> ReadSheetExcel:
        rd = ExcelLoadOpenPyxl(file_excel)
        return cls(rd)

    @classmethod
    def create_load_pandas(cls, file_excel: str | BytesIO) -> ReadSheetExcel:
        rd = ExcelLoadPandas(file_excel)
        return cls(rd)
